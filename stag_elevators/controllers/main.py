from collections import defaultdict
import os
from requests import Response
from odoo import fields, http
from odoo.http import request
import calendar
from datetime import datetime, date
from werkzeug.exceptions import NotFound
from odoo.exceptions import AccessDenied, UserError, ValidationError
import io
import csv
import base64
from datetime import datetime
from openpyxl import load_workbook
import re
import json
from datetime import date, timedelta






from odoo import http, _
from odoo.http import request
from odoo.addons.web.controllers.home import Home

class CustomLoginController(http.Controller):

    @http.route('/web/login', type='http', auth="public", website=True, csrf=True)
    def web_login(self, redirect=None, **kw):
        login = kw.get('login')
        password = kw.get('password')
        db = kw.get('db') or request.session.db  # use selected DB

        # Check if user exists and is inactive
        if login:
            user = request.env['res.users'].sudo().with_env(request.env).search([('login', '=', login)], limit=1)
            if not user:

                user = request.env['res.users'].sudo().search([
                    ('login', '=', login),
                    ('active', '=', False)   # only inactive users
                ], limit=1)
            if user and not user.active:
                # Render custom login template with error
                return request.render('stag_elevators.custom_login_template', {
                    'error': _('User is deactivated, kindly contact admin'),
                    'login': login or '',
                })

        # If user is active (or login empty), delegate to Odoo's original login
        return Home().web_login(redirect=redirect, **kw)




    




class WebRedirect(http.Controller):

    @http.route('/', type='http', auth='public', website=True)
    def index_redirect(self, **kw):
        if not request.session.uid:
            return request.redirect('/web/login')
        else:
            return request.redirect('/Dashboard')
        
    

class LeadDashboardController(http.Controller):

    @http.route('/Dashboard', type='http', auth='user', website=True)
    def dashboard_page(self, **kwargs):
        sales=True
        crm=True
        production=True
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.client'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            crm=False
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'lead.followup'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            sales=False

        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'production.client'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            production=False

        # access_rule_admin = request.env['ir.model.access'].sudo().search([
        #         ('model_id.model', '=', 'ir.model.access'),
        #         ('group_id', '=', my_groups.id),
        #         ('perm_read', '=', True)
        #     ], limit=1)
        if  my_groups.is_admin:
            today = date.today()
            first_day = today.replace(day=1)
            last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)

            user = request.env.user
            total_leads = request.env['crm.lead'].sudo().search_count([])

            # Safely fetch the "Won" stage by name
            won_stage = request.env['crm.stage'].sudo().search([('name', 'ilike', 'won')], limit=1)

            # Initialize count_won
            count_won = 0
            if won_stage:
                count_won = request.env['lead.followup'].sudo().search_count([
                    ('stage_id', '=', won_stage.id),
                    ('next_followup_date', '>=', first_day),
                    ('next_followup_date', '<=', last_day),
                ])
                
                

            # Fetch all CRM stages and followup counts
            stages = request.env['crm.stage'].sudo().search([], order="sequence asc")
            stage_followup_counts = []
            stage_crm_counts = []
            stage_production_counts = []

            for stage in stages:
                if stage.stage == 'sales':
                    count = request.env['crm.lead'].sudo().search_count([
                        ('stage_id', '=', stage.id)
                    ])
                    stage_followup_counts.append({
                        'stage': stage.name,
                        'count': count,
                        'id': stage.id,
                    })
                elif stage.stage == 'crm':
                    count = request.env['crm.client'].sudo().search_count([
                        ('stage_id', '=', stage.id)
                    ])
                    stage_crm_counts.append({
                        'stage': stage.name,
                        'count': count,
                        'id': stage.id,
                    })
                elif stage.stage == 'production':
                    count = request.env['production.client'].sudo().search_count([
                        ('stage_id', '=', stage.id)
                    ])
                    stage_production_counts.append({
                        'stage': stage.name,
                        'count': count,
                        'id': stage.id,
                    })


            # You can add filtering later for pending if needed
            pending_followups = request.env['lead.followup'].sudo().search([('lead_id.stage_id.name', '!=', 'Won'),('lead_id.stage_id.name', '!=', 'Lost')])

            amc_renewals = request.env['crm.client'].sudo().search([
                
                ('renew', '=', True)
            ])
            dispatch = request.env['production.client'].sudo().search([('material_dispatch_date','!=',False)])
            


            return request.render('stag_elevators.dashboard_template', {
                'total_leads': total_leads,
                'count_won': count_won,
                'stage_lead_counts': stage_followup_counts,
                'stage_crm_counts':stage_crm_counts,
                'stage_production_counts':stage_production_counts,
                'pending_followups': pending_followups,
                'user_name': user.name,
                'user_image': user.image_128.decode() if user.image_128 else '',
                'amc_renewals':amc_renewals,
                'crm':crm,
                'sales':sales,
                'dispatch':dispatch,
                'production':production,
            })

        today = date.today()
        first_day = today.replace(day=1)
        last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        user = request.env.user
        total_leads = request.env['crm.lead'].sudo().search_count([('user_id', '=', user.id)])

        # Safely fetch the "Won" stage by name
        won_stage = request.env['crm.stage'].sudo().search([('name', 'ilike', 'won')], limit=1)

        # Initialize count_won
        count_won = 0
        if won_stage:
            count_won = request.env['lead.followup'].sudo().search_count([
                ('stage_id', '=', won_stage.id),
                ('next_followup_date', '>=', first_day),
                ('next_followup_date', '<=', last_day),
                ('lead_id.user_id','=',user.id),
                
            ])
            if crm:
                    count_won = request.env['crm.client'].sudo().search_count([
                    ('assign_to','=',user.id)
                ])

        # Fetch all CRM stages and followup counts
        stages = request.env['crm.stage'].sudo().search([], order="sequence asc")
        stage_followup_counts = []
        stage_crm_counts=[]
        stage_production_counts = []

        for stage in stages:
            if stage.stage == 'sales':
                count = request.env['crm.lead'].sudo().search_count([
                    ('stage_id', '=', stage.id),
                    ('user_id', '=', user.id)
                    
                ])
                stage_followup_counts.append({
                    'stage': stage.name,
                    'count': count,
                    'id': stage.id,
                })
            elif stage.stage == 'crm':
                count = request.env['crm.client'].sudo().search_count([
                    ('stage_id', '=', stage.id),
                    ('assign_to', '=', request.env.user.id),
                ])
                stage_crm_counts.append({
                    'stage': stage.name,
                    'count': count,
                    'id': stage.id,
                })
            elif stage.stage == 'production':
                    count = request.env['production.client'].sudo().search_count([
                        ('stage_id', '=', stage.id),
                        ('assign_to', '=', request.env.user.id),
                    ])
                    stage_production_counts.append({
                        'stage': stage.name,
                        'count': count,
                        'id': stage.id,
                    })
            

        # You can add filtering later for pending if needed
        pending_followups = request.env['lead.followup'].sudo().search([('lead_id.user_id','=',user.id),('lead_id.stage_id.name', '!=', 'Lost'),('lead_id.stage_id.name', '!=', 'Won')])

        amc_renewals = request.env['crm.client'].sudo().search([
            ('assign_to', '=', request.env.user.id),
            ('renew', '=', True)
        ])
        dispatch = request.env['production.client'].sudo().search([('material_dispatch_date','!=',False),('assign_to', '=', request.env.user.id),])


        return request.render('stag_elevators.dashboard_template', {
            'total_leads': total_leads,
            'count_won': count_won,
            'stage_lead_counts': stage_followup_counts,
            'stage_crm_counts':stage_crm_counts,
            'pending_followups': pending_followups,
            'stage_production_counts':stage_production_counts,
            'user_name': user.name,
            'user_image': user.image_128.decode() if user.image_128 else '',
            'amc_renewals':amc_renewals,
            'crm':crm,
            'sales':sales,
            'dispatch':dispatch,
            'production':production,
        })


    

class CityWebsiteController(http.Controller):

    # @http.route('/city/list', type='http', auth='public', website=True)
    # def city_list(self, **kwargs):
    #     search_term = kwargs.get('search', '').strip()
    #     domain = []

    #     if search_term:
    #         domain = ['|',
    #                 ('name', 'ilike', search_term),
    #                 ('status', 'ilike', search_term)]

    #     cities = request.env['res.city'].sudo().search(domain)
    #     return request.render('stag_elevators.city_list_template', {
    #         'cities': cities,
    #         'search_term': search_term
    #     })

    # @http.route('/city/create', type='http', auth='public', website=True, csrf=True)
    # def city_create(self, **post):
    #     if request.httprequest.method == 'POST':
    #         name = post.get('city_name')
    #         status = post.get('status')
    #         if name:
    #             request.env['res.city'].sudo().create({
    #                 'name': name,
    #                 'status': status or 'active',
    #             })
    #             return request.redirect('/city/list')
    #     return request.render('stag_elevators.city_create_template')

    # @http.route('/city/edit/<int:city_id>', type='http', auth='public', website=True, csrf=True)
    # def city_edit(self, city_id, **post):
    #     city = request.env['res.city'].sudo().browse(city_id)
    #     if request.httprequest.method == 'POST':
    #         city.name = post.get('city_name')
    #         city.status = post.get('status') or 'active'
    #         return request.redirect('/city/list')
    #     return request.render('stag_elevators.city_edit_template', {'city': city})

    # @http.route('/city/delete/<int:city_id>', type='http', auth='public', website=True, csrf=True)
    # def city_delete(self, city_id, **kwargs):
    #     city = request.env['res.city'].sudo().browse(city_id)
    #     if city:
    #         city.unlink()
    #     return request.redirect('/city/list')

    @http.route('/city/list', type='http', auth='user', website=True)
    def city_list(self, **kwargs):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.city'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view cities.'
            })
        cities = request.env['res.city'].sudo().search([])
        return request.render('stag_elevators.city_list_template', {
            'cities': cities
        })
    
    @http.route('/city/create', type='http', auth='user', website=True, csrf=True)
    def city_create(self, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.city'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view cities.'
            })
        name= post.get('city_name')
        if name:
                cityy = request.env['res.city'].sudo().search([('name','=',name)],limit=1)
                if cityy:
                    return request.redirect(f'/city/create?error=city_duplicate')
        if request.httprequest.method == 'POST':
            name = post.get('city_name')
            status = post.get('status')
            if name:
                request.env['res.city'].sudo().create({
                    'name': name,
                    'status': status or 'active',
                })
                return request.redirect('/city/list')
        return request.render('stag_elevators.city_create_template')


    @http.route('/city/edit/<int:city_id>', type='http', auth='user', website=True, csrf=True)
    def city_edit(self, city_id, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.city'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view cities.'
            })
        city = request.env['res.city'].sudo().browse(city_id)
        if not city.exists():
            return request.not_found()

        if request.httprequest.method == 'POST':
            city.name = post.get('city_name')
            city.status = post.get('status') or 'active'
            return request.redirect('/city/list')

        return request.render('stag_elevators.city_edit_template', {'city': city})

    @http.route('/city/delete/<int:city_id>', type='http', auth='user', website=True, csrf=True)
    def city_delete(self, city_id, **kwargs):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.city'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view cities.'
            })
        city = request.env['res.city'].sudo().browse(city_id)
        if city.exists():
            city.unlink()
        return request.redirect('/city/list')
    

class CRMStageController(http.Controller):

    @http.route('/lead_stage/list', type='http', auth='user', website=True)
    def lead_stage_list(self, **kwargs):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.stage'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        stages = request.env['crm.stage'].sudo().search([], order='sequence DESC')
        return request.render('stag_elevators.crm_stage_list_template', {
            'stages': stages
        })
    
    @http.route('/lead_stage/create', type='http', auth='user', website=True)
    def lead_stage_create_form(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.stage'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        return request.render('stag_elevators.crm_stage_create_template', {})

    @http.route('/lead_stage/submit', type='http', auth='user', website=True, csrf=True)
    def lead_stage_submit(self, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.stage'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        name = post.get('name')
        sequence = int(post.get('sequence') or 0)
        description = post.get('description')
        is_won = post.get('is_won') == 'yes'
        status = post.get('status')
    
        stage_type = post.get('stage')
        if name:
            namee = request.env['crm.stage'].sudo().search([('name','=',name),('stage','=',stage_type)],limit=1)
            if namee:
                return request.redirect(f'/lead_stage/create?error=stage_duplicate')

        # Create the CRM Stage
        request.env['crm.stage'].sudo().create({
            'name': name,
            'sequence': sequence,
            'description': description,
            'is_won': is_won,
            'status':status,

            'stage':stage_type,
        })

        return request.redirect('/lead_stage/list')
    
    @http.route('/lead_stage/edit/<int:stage_id>', type='http', auth='user', website=True)
    def lead_stage_edit_form(self, stage_id, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.stage'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        stage = request.env['crm.stage'].sudo().browse(stage_id)
        if not stage.exists():
            raise NotFound("Lead Stage not found")
        return request.render('stag_elevators.crm_stage_edit_template', {'stage': stage})

    @http.route('/lead_stage/update', type='http', auth='user', website=True, csrf=True)
    def lead_stage_update(self, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.stage'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        stage_id = int(post.get('stage_id'))
        name =post.get('name')
        # is_crm = post.get('is_crm')
        stage_type = post.get('stage')
        if name:
            namee = request.env['crm.stage'].sudo().search([('name','=',name),('stage','=',stage_type),('id', '!=', stage_id)],limit=1)
            if namee:
                return request.redirect(f'/lead_stage/create?error=stage_duplicate')
        stage = request.env['crm.stage'].sudo().browse(stage_id)
        
        if stage.exists():
            stage.write({
                'name': post.get('name'),
                'sequence': int(post.get('sequence') or 0),
                'description': post.get('description'),
                'is_won': post.get('is_won') == 'yes',
                'status': post.get('status'),

                'stage':stage_type
            })
        return request.redirect('/lead_stage/list')
    
    @http.route('/lead_stage/delete/<int:stage_id>', type='http', auth='user', website=True)
    def lead_stage_delete(self, stage_id, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.stage'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        stage = request.env['crm.stage'].sudo().browse(stage_id)
        if stage.exists():
            stage.unlink()
        return request.redirect('/lead_stage/list')
    
class UserRole(http.Controller):

    @http.route('/user_permissions', type='http', auth='user', website=True)
    def user_permission_role_list(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'ir.model.access'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        roles = request.env['res.groups'].sudo().search([('ismy', '=', True)])
        return request.render('stag_elevators.user_permission_role_list_template', {
            'roles': roles
        })
    
    @http.route('/user_permission/permissions/<int:role_id>', type='http', auth='user', website=True)
    def user_permission_role_edit(self, role_id, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'ir.model.access'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        role = request.env['res.groups'].sudo().browse(role_id)
        def has_access(group, model_name, perm='read'):
            model = request.env['ir.model'].sudo().search([('model', '=', model_name)], limit=1)
            if not model:
                return False
            access = request.env['ir.model.access'].sudo().search([
                ('group_id', '=', group.id),
                ('model_id', '=', model.id)
            ], limit=1)
            if not access:
                return False
            return getattr(access, f'perm_{perm}', False)
        return request.render('stag_elevators.user_permission_template_edit', {
            'role': role,
            'has_access': has_access,
        })
        

    @http.route('/user_permission/list', type='http', auth='public', website=True)
    def user_permission_list(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.groups'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        roles = request.env['res.groups'].sudo().search([('ismy', '=', True)])
        return request.render('stag_elevators.user_permission_list_template', {
            'roles': roles
        })

    @http.route('/user_permission/create', type='http', auth='public', website=True)
    def user_role_create(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.groups'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        return request.render('stag_elevators.user_role_create_template')
    
    @http.route('/user_permission/edit/<int:role_id>', type='http', auth='public', website=True)
    def user_permission_edit(self, role_id, **kw):
        # Check if current user has "ismy=True" group and read access
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
            ('model_id.model', '=', 'res.groups'),
            ('group_id', 'in', my_groups.ids),
            ('perm_read', '=', True)
        ], limit=1)

        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })

        # Fetch the role (group) being edited
        role = request.env['res.groups'].sudo().browse(role_id)
        if not role.exists():
            return request.render('stag_elevators.access_denied_template', {
                'message': 'Role not found.'
            })

        return request.render('stag_elevators.user_permission_edit_template', {
            'role': role
        })
    
    @http.route('/user_permission/update', type='http', auth='public', methods=['POST'], website=True, csrf=True)
    def user_permission_update(self, **post):
        role_id = int(post.get('role_id'))
        role = request.env['res.groups'].sudo().browse(role_id)
        if role.exists():
            role.sudo().write({
                'name': post.get('name'),
                'is_admin':bool(post.get('is_admin')),
            })
        return request.redirect('/user_permission/list')


    @http.route('/user_permission/save', type='http', auth='public', website=True, methods=['POST'])
    def user_role_save(self, **post):
        name = post.get('name')
        is_admin = bool(post.get('is_admin'))
        if not name:
            raise UserError("Role name is required.")
        if name:
                namee = request.env['res.groups'].sudo().search([('name','=',name)],limit=1)
                if namee:
                    return request.redirect(f'/user_permission/create?error=role_duplicate')
        request.env['res.groups'].sudo().create({
        'name': name,
        'ismy': True,
        'is_admin':is_admin,   
    })
        return request.redirect('/user_permission/list')
    
    @http.route('/user_permission/delete/<int:role_id>', type='http', auth='user', website=True)
    def delete_user_permission(self, role_id, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.groups'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        role = request.env['res.groups'].sudo().browse(role_id)
        if role.exists():
            role.unlink()
        return request.redirect('/user_permission/list')
    
class UserController(http.Controller):

    @http.route('/users', type='http', auth='user', website=True)
    def user_list(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.users'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        # users = request.env['res.users'].sudo().search([])
        users = request.env['res.users'].with_context(active_test=False).sudo().search([
            ('groups_id.ismy', '=', True),
        ])
        return request.render('stag_elevators.user_list_template', {
            'users': users
        })

    @http.route('/users/create', type='http', auth='user', website=True)
    def user_create_form(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.users'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        cities = request.env['res.city'].sudo().search([])
        users = request.env['res.groups'].sudo().search([('ismy', '=', True)])
        return request.render('stag_elevators.user_form_template', {
            'cities': cities,
            'users':users,
        })

    @http.route('/users/submit', type='http', auth='user', website=True, csrf=True, methods=['POST'])
    def user_submit(self, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.users'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        name1=post.get('user_name')
        name = post.get('username')
        password = post.get('password')
        # role_id = int(post.get('role_id')) if post.get('role_id') else False
        location = post.get('location')
        status = post.get('status')
        if name:
            # namee = request.env['res.users'].sudo().search([('login','=',name)],limit=1)
            namee = request.env['res.users'].with_context(active_test=False).sudo().search(
                [('login', '=', name)], limit=1
            )
            if namee:
                return request.redirect(f'/users/create?error=user_duplicate')

        partner = request.env['res.partner'].sudo().create({'name': name})

        # user_vals = {
        #     'name': name1,
        #     'login': name,
        #     'password': password,
        #     'active': True if status == 'active' else False,
        #     'partner_id': partner.id,
        #     'role_id': role_id,   
        # }

        # user = request.env['res.users'].sudo().create(user_vals)

        role_id = post.get('role_id')
        try:
            group = request.env['res.groups'].sudo().browse(int(role_id))
        except ValueError:
            group = request.env['res.groups'].sudo().search([('name', '=', role_id)], limit=1)

        if not group:
            print("not found")
        basic_group = request.env.ref("base.group_user")

        user_vals = {
                    'name': name1,
                    'login': name,
                    'password': password,
                    'active': True if status == 'active' else False,
                    'partner_id': partner.id,
                    'groups_id': [(6, 0, [basic_group.id, group.id])],
                    'location':location,
                    'signature':password,
                }

        user = request.env['res.users'].sudo().create(user_vals)

        # assign group from role
        # if user.role_id and user.role_id.group_id:
        #     user.sudo().write({
        #         'groups_id': [(6, 0, [user.role_id.group_id.id])]
        #     })

        return request.redirect('/users')
    
    # Edit page
    @http.route('/users/edit/<int:user_id>', type='http', auth='user', website=True)
    def edit_user(self, user_id):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.users'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        user = request.env['res.users'].sudo().browse(user_id)
        cities = request.env['res.city'].sudo().search([])
        roles = request.env['res.groups'].sudo().search([('ismy', '=', True)])
        return request.render('stag_elevators.user_edit_template', {
            'user': user,
            'cities': cities,
            'roles': roles,
        })
    
    @http.route('/users/update/<int:user_id>', type='http', auth='user', website=True, csrf=True, methods=['POST'])
    def update_user(self, user_id, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.users'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        user = request.env['res.users'].sudo().browse(user_id)

        if not user.exists():
            return request.not_found()

        vals = {
            'name': post.get('user_name'),
            # 'login': post.get('username'),
            'active': True if post.get('status') == 'active' else False,
            'location':post.get('location'),
            'signature':post.get('password'),
        }
        location = post.get('location')

        if post.get('password'):
            vals['password'] = post.get('password')


        role_id = post.get('role_id')
        if role_id:
            try:
                group = request.env['res.groups'].sudo().browse(int(role_id))
            except ValueError:
                group = request.env['res.groups'].sudo().search([('name', '=', role_id)], limit=1)

            if group:
                basic_group = request.env.ref("base.group_user")
                vals['groups_id'] = [(6, 0, [basic_group.id, group.id])]


        if user.partner_id:
            user.partner_id.sudo().write({'name': post.get('username')})

        user.sudo().write(vals)
        user.location = int(location)

        return request.redirect('/users')


    # Delete route
    @http.route('/users/delete/<int:user_id>', type='http', auth='user', website=True)
    def delete_user(self, user_id):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'res.users'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        user = request.env['res.users'].sudo().browse(user_id)
        if user:
            user.unlink()
        return request.redirect('/users')

    
class OpportunityController(http.Controller):

    @http.route('/opportunities', type='http', auth='user', website=True)
    def list_opportunities(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        
        # access_rule_admin = request.env['ir.model.access'].sudo().search([
        #         ('model_id.model', '=', 'ir.model.access'),
        #         ('group_id', '=', my_groups.id),
        #         ('perm_read', '=', True)
        #     ], limit=1)
        if my_groups.is_admin:
            leads = request.env['crm.lead'].sudo().search([], order='id desc')
            
            # users = request.env['res.users'].sudo().search([])  # Fetch all salespersons
            my_groups = request.env['res.groups'].sudo().search([('ismy', '=', True)])
            users_in_my_groups = request.env['res.users'].sudo().search([
                ('active', '=', True),
                ('groups_id', 'in', my_groups.ids)
            ])
            final_users = []
            for user in users_in_my_groups:
                access_rule = request.env['ir.model.access'].sudo().search([
                    ('model_id.model', '=', 'crm.lead'),
                    ('group_id', 'in', user.groups_id.ids),
                    ('perm_read', '=', True)
                ], limit=1)
                access_rule_admin = request.env['ir.model.access'].sudo().search([
                    ('model_id.model', '=', 'ir.model.access'),
                    ('group_id', 'in', user.groups_id.ids),
                    ('perm_read', '=', True)
                ], limit=1)
                
                if access_rule:
                    if not access_rule_admin:
                        final_users.append(user)
            users = request.env['res.users'].browse([u.id for u in final_users])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
        else:
            request.env.cr.rollback()  # resets the transaction
            user = request.env.user
            leads = request.env['crm.lead'].sudo().search([ '|',
    ('user_id', '=', user.id),
    ('create_uid', '=', user.id) ], order='create_date desc')
            my_groups = request.env['res.groups'].sudo().search([('ismy', '=', True)])
            users_in_my_groups = request.env['res.users'].sudo().search([
                ('active', '=', True),
                ('groups_id', 'in', my_groups.ids)
            ])
            final_users = []
            for user in users_in_my_groups:
                access_rule = request.env['ir.model.access'].sudo().search([
                    ('model_id.model', '=', 'crm.lead'),
                    ('group_id', 'in', user.groups_id.ids),
                    ('perm_read', '=', True)
                ], limit=1)
                access_rule_admin = request.env['ir.model.access'].sudo().search([
                    ('model_id.model', '=', 'ir.model.access'),
                    ('group_id', 'in', user.groups_id.ids),
                    ('perm_read', '=', True)
                ], limit=1)
                
                if access_rule:
                    if not access_rule_admin:
                        final_users.append(user)
            users = request.env['res.users'].browse([u.id for u in final_users])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
    
    @http.route('/opportunity/assign/single', type='http', auth='user', methods=['POST'], csrf=False)
    def assign_single(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        data = json.loads(request.httprequest.data)
        lead_id = data.get('lead_id')
        user_id = data.get('user_id')
        if lead_id and user_id:
            lead = request.env['crm.lead'].sudo().browse(int(id))
            lead.user_id = int(user_id)
            return json.dumps({'success': True})
        return json.dumps({'success': False})

    @http.route('/opportunity/assign/bulk', type='http', auth='user', methods=['POST'], csrf=False)
    def assign_bulk(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        if my_groups.is_admin:
            data = json.loads(request.httprequest.data)
            assignments = data.get('assignments', [])
            for item in assignments:
                lead_id = item.get('lead_id')
                user_id = item.get('user_id')
                if lead_id and user_id:
                    lead = request.env['crm.lead'].sudo().browse(int(lead_id))
                    lead.user_id = int(user_id)
            return json.dumps({'success': True})
        else:
            return json.dumps({'success': False})
    
    # @http.route('/assign_selected_leads', type='http', auth='user', website=True, methods=['POST'])
    # def assign_selected_leads(self, **kw):
    #     assignments = kw.get('assignments')
    #     if not assignments:
    #         return "No assignments provided", 400

    #     assignments = json.loads(assignments)  
        
    #     for assignment in assignments:
    #         lead_id = assignment.get('lead_id')
    #         salesperson_id = assignment.get('salesperson_id')
            
    #         lead = request.env['crm.lead'].sudo().browse(lead_id)
    #         salesperson = request.env['res.users'].sudo().browse(salesperson_id)

    #         if lead.exists() and salesperson.exists():
    #             lead.write({'user_id': salesperson.id})  # Assign salesperson to the lead

    #     return "Leads assigned successfully"
    
    
    @http.route('/opportunities/new', type='http', auth='user', website=True)
    def list_opportunities_new(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        if my_groups.is_admin:
            new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'New')], limit=1)
            
            if new_stage:
                leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),], order='create_date desc')
            else:
                leads = request.env['crm.lead'].sudo().search([])
            users = request.env['res.users'].sudo().search([])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
        user = request.env.user
        new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'New')], limit=1)
            
        if new_stage:
            leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),('user_id','=',user.id)], order='create_date desc')
        else:
            leads = request.env['crm.lead'].sudo().search([])
        users = request.env['res.users'].sudo().search([])
        return request.render('stag_elevators.opportunity_list_template', {
            'leads': leads,
            'users': users,
        })
        
    
    @http.route('/opportunities/qualified', type='http', auth='user', website=True)
    def list_opportunities_qualified(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        if my_groups.is_admin:
            new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Qualified')], limit=1)
            
            if new_stage:
                leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id)], order='create_date desc')
            else:
                leads = request.env['crm.lead'].sudo().search([])
            users = request.env['res.users'].sudo().search([])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
        else:
            user = request.env.user
            new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Qualified')], limit=1)
            
            if new_stage:
                leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),('user_id','=',user.id)], order='create_date desc')
            else:
                leads = request.env['crm.lead'].sudo().search([])
            users = request.env['res.users'].sudo().search([])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })

    
    @http.route('/opportunities/Proposition', type='http', auth='user', website=True)
    def list_opportunities_proposition(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        if my_groups.is_admin:
            new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Proposition')], limit=1)
            
            if new_stage:
                leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id)], order='create_date desc')
            else:
                leads = request.env['crm.lead'].sudo().search([])
            users = request.env['res.users'].sudo().search([])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
        else:
            user = request.env.user
            new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Proposition')], limit=1)
            
            if new_stage:
                leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),('user_id','=',user.id)], order='create_date desc')
            else:
                leads = request.env['crm.lead'].sudo().search([])
            users = request.env['res.users'].sudo().search([])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
        
    @http.route('/opportunities/hot', type='http', auth='user', website=True)
    def list_opportunities_hot(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        if my_groups.is_admin:
            new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Hot')], limit=1)
            
            if new_stage:
                leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),], order='create_date desc')
            else:
                leads = request.env['crm.lead'].sudo().search([])
            users = request.env['res.users'].sudo().search([])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
        user = request.env.user
        new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Hot')], limit=1)
            
        if new_stage:
            leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),('user_id','=',user.id)], order='create_date desc')
        else:
            leads = request.env['crm.lead'].sudo().search([])
        users = request.env['res.users'].sudo().search([])
        return request.render('stag_elevators.opportunity_list_template', {
            'leads': leads,
            'users': users,
        })
    
    from odoo import http
from odoo.http import request

class OpportunityController(http.Controller):

    @http.route('/opportunities/stage', type='http', auth='user', website=True)
    def list_opportunities(self, stage=None, **kw):
        # Check user groups & access
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
            ('model_id.model', '=', 'crm.lead'),
            ('group_id', '=', my_groups.id),
            ('perm_read', '=', True)
        ], limit=1)

        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })

        # Stage filter (case-insensitive)
        stage_domain = []
        if stage:
            new_stage = request.env['crm.stage'].sudo().search([('name', 'ilike', stage),('stage','=','sales')], limit=1)
            if new_stage:
                stage_domain = [('stage_id', '=', new_stage.id)]

        # If admin → all leads in stage
        if my_groups.is_admin:
            leads = request.env['crm.lead'].sudo().search(stage_domain, order='create_date desc')
        else:
            user = request.env.user
            leads = request.env['crm.lead'].sudo().search(stage_domain + [('user_id', '=', user.id)], order='create_date desc')

        # users = request.env['res.users'].sudo().search([])
        final_users = []
        users_in_my_groups = request.env['res.users'].sudo().search([
                ('active', '=', True),
                # ('groups_id', 'in', my_groups.ids)
            ])
        for user in users_in_my_groups:
            access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', 'in', user.groups_id.ids),
                ('perm_read', '=', True)
            ], limit=1)
            access_rule_admin = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'ir.model.access'),
                ('group_id', 'in', user.groups_id.ids),
                ('perm_read', '=', True)
            ], limit=1)
            
            if access_rule:
                if not access_rule_admin:
                    final_users.append(user)
        users = request.env['res.users'].browse([u.id for u in final_users])

        return request.render('stag_elevators.opportunity_list_template', {
            'leads': leads,
            'users': users,
            'stage': stage,
        })

    
    @http.route('/opportunities/warm', type='http', auth='user', website=True)
    def list_opportunities_warm(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        if my_groups.is_admin:
            new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Warm')], limit=1)
            
            if new_stage:
                leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),], order='create_date desc')
            else:
                leads = request.env['crm.lead'].sudo().search([])
            users = request.env['res.users'].sudo().search([])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
        user = request.env.user
        new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Warm')], limit=1)
            
        if new_stage:
            leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),('user_id','=',user.id)], order='create_date desc')
        else:
            leads = request.env['crm.lead'].sudo().search([])
        users = request.env['res.users'].sudo().search([])
        return request.render('stag_elevators.opportunity_list_template', {
            'leads': leads,
            'users': users,
        })
    
    @http.route('/opportunities/cold', type='http', auth='user', website=True)
    def list_opportunities_cold(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        if my_groups.is_admin:
            new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Cold')], limit=1)
            
            if new_stage:
                leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),], order='create_date desc')
            else:
                leads = request.env['crm.lead'].sudo().search([])
            users = request.env['res.users'].sudo().search([])
            return request.render('stag_elevators.opportunity_list_template', {
                'leads': leads,
                'users': users,
            })
        user = request.env.user
        new_stage = request.env['crm.stage'].sudo().search([('name', '=', 'Cold')], limit=1)
            
        if new_stage:
            leads = request.env['crm.lead'].sudo().search([('stage_id', '=', new_stage.id),('user_id','=',user.id)], order='create_date desc')
        else:
            leads = request.env['crm.lead'].sudo().search([])
        users = request.env['res.users'].sudo().search([])
        return request.render('stag_elevators.opportunity_list_template', {
            'leads': leads,
            'users': users,
        })
    
    @http.route('/assign/salesperson', type='http', auth='user', methods=['POST'], website=True)
    def assign_salesperson(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        assigned_to = kw.get('assigned_to')
        leads = request.env['crm.lead'].sudo().search([])  # Fetch all leads
        for lead in leads:
            lead.user_id = assigned_to  # Assign the selected salesperson
        return request.redirect('/opportunities')

    # @http.route('/opportunities', type='http', auth='user', website=True)
    # def opportunity_list(self, **kw):
    #     leads = request.env['crm.lead'].sudo().search([], order='create_date desc')
    #     seen = defaultdict(list)
    #     duplicates_map = {}

    #     for lead in leads:
    #         key = (lead.name.strip().lower() if lead.name else '', lead.phone or '')
    #         seen[key].append(lead.id)

    #     # Mark duplicates in dictionary
    #     for ids in seen.values():
    #         if len(ids) > 1:
    #             for lead_id in ids:
    #                 duplicates_map[lead_id] = True

    #     return request.render('stag_elevators.opportunity_list_template', {'leads': leads,
    #                                                                       'duplicates_map': duplicates_map})
    
    @http.route('/opportunities/create', type='http', auth='user', website=True)
    def create_opportunity_form(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        # users = request.env['res.users'].sudo().search([])
        users_in_my_groups = request.env['res.users'].sudo().search([
                ('active', '=', True),
                # ('groups_id', 'in', my_groups.ids)
            ])
        final_users = []
        for user in users_in_my_groups:
            access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', 'in', user.groups_id.ids),
                ('perm_read', '=', True)
            ], limit=1)
            
            if access_rule:
                final_users.append(user)
        users = request.env['res.users'].browse([u.id for u in final_users])
        locations = request.env['res.city'].sudo().search([])  # Or static list if not using res.city
        return request.render('stag_elevators.opportunity_create_template', {
            'users': users,
            'locations': locations,
        })
    
    # @http.route('/opportunities/check_duplicate', type='http', auth='user', csrf=True)
    # def check_duplicate_lead(self, **post):
    #     phone = (post.get("phone") or "").strip()
    #     if not phone:
    #         return request.make_response(json.dumps({"exists": False}), headers=[("Content-Type", "application/json")])

    #     lead = request.env['crm.lead'].sudo().search([('phone', '=', phone)], limit=1)
    #     if lead:
    #         return request.make_response(json.dumps({"exists": True, "lead_id": lead.id, "phone": phone}), headers=[("Content-Type", "application/json")])
    #     return request.make_response(json.dumps({"exists": False}), headers=[("Content-Type", "application/json")])



    
    @http.route('/opportunities/submit', type='http', auth='user', website=True, csrf=True)
    def create_opportunity_submit(self, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        name = post.get('name', '').strip()
        phone = str(post.get('phone', '') or '').strip()
        email = post.get('email', '').strip()

        if phone:
            namee = request.env['crm.lead'].sudo().search([('phone','=',phone)],limit=1)
            if namee:
                return request.redirect(f'/opportunities/create?error=lead_duplicate')

        # Name validation
        if not name:
            raise UserError("Name is required.")

        # Phone validation (basic: digits and length between 7 to 15)
        if phone and not (phone.isdigit() and 7 <= len(phone) <= 15):
            raise UserError(f"Invalid phone number: {phone}")
        
        

        # Email validation using regex
        email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        if email and not re.match(email_regex, email):
            raise UserError(f"Invalid email address: {email}")
        
        stage = request.env['crm.stage'].search([('name', '=', 'New')], limit=1)
        if my_groups.is_admin:
        # Create lead if all validations pass
            request.env['crm.lead'].sudo().create({
                'name': name,
                'phone': phone,
                # 'x_alternate_mobile': post.get('alt_phone'),
                'email_from': email,
                'city': post.get('city'),
                'user_id': int(post.get('user_id')) if post.get('user_id') else request.env.user.id,
                'description': post.get('description'),
                'stage_id': stage.id,
                
            })
        else:
            request.env['crm.lead'].sudo().create({
                'name': name,
                'phone': phone,
                # 'x_alternate_mobile': post.get('alt_phone'),
                'email_from': email,
                'city': post.get('city'),
                'user_id':  request.env.user.id,
                'description': post.get('description'),
                'stage_id': stage.id,
            })

        return request.redirect('/opportunities')
    
    @http.route('/opportunity/transfer/<int:lead_id>', type='http', auth='user', website=True)
    def transfer_lead_form(self, lead_id, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        if my_groups.is_admin:
            lead = request.env['crm.lead'].sudo().browse(lead_id)
            users = request.env['res.users'].sudo().search([])
            my_groups = request.env['res.groups'].sudo().search([('ismy', '=', True)])
            users_in_my_groups = request.env['res.users'].sudo().search([
                ('active', '=', True),
                
            ])
            final_users = []
            for user in users_in_my_groups:
                access_rule = request.env['ir.model.access'].sudo().search([
                    ('model_id.model', '=', 'crm.lead'),
                    ('group_id', 'in', user.groups_id.ids),
                    ('perm_read', '=', True)
                ], limit=1)
                
                if access_rule:
                    final_users.append(user)
            users = request.env['res.users'].browse([u.id for u in final_users])
            locations = request.env['res.city'].sudo().search([])
            
            # Count leads per user
            for user in users:
                user.lead_count = request.env['crm.lead'].sudo().search_count([('user_id', '=', user.id)])

            return request.render('stag_elevators.transfer_lead_template', {
                'lead': lead,
                'users': users,
                'locations': locations,
            })
        else:
            return request.redirect('/opportunities')

    @http.route('/opportunity/transfer/submit', type='http', auth='user', website=True, csrf=True)
    def transfer_lead_submit(self, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        
        lead_id = int(post.get('lead_id'))
        lead = request.env['crm.lead'].sudo().browse(lead_id)

        lead.write({

            'city': post.get('city'),
            'user_id': int(post.get('user_id')),
        })

        return request.redirect('/opportunities')
    

class LeadImportController(http.Controller):

    @http.route('/download/lead_template', type='http', auth='user')
    def download_lead_template(self, **kw):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        # Get absolute path of the module
        module_path = os.path.dirname(os.path.abspath(__file__))  # controllers folder
        file_path = os.path.join(module_path, "..", "static", "files", "lead_template.xlsx")
        file_path = os.path.normpath(file_path)  # normalize path for Windows/Linux

        if not os.path.exists(file_path):
            return request.not_found()

        with open(file_path, 'rb') as f:
            file_data = f.read()

        return request.make_response(
            file_data,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="Lead_Template.xlsx";')
            ]
        )
    
    @http.route('/import/leads', type='http', auth='user', website=True)
    def import_leads_form(self, **kwargs):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        return request.render('stag_elevators.lead_import_template')
    

# =============== working model ===============



    # @http.route('/import/leads/submit', type='http', auth='user', website=True, csrf=False)
    # def import_leads_submit(self, **post):
    #     my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
    #     access_rule = request.env['ir.model.access'].sudo().search([
    #             ('model_id.model', '=', 'crm.lead'),
    #             ('group_id', '=', my_groups.id),
    #             ('perm_read', '=', True)
    #         ], limit=1)
    #     if not access_rule:
    #         return request.render('stag_elevators.access_denied_template', {
    #             'message': 'You do not have permission to view.'
    #         })
    #     file = post.get('file')
    #     if not file:
    #         return request.render('stag_elevators.import_error_template', {
    #             'error': 'No file uploaded.'
    #         })

    #     leads = request.env['crm.lead'].sudo()
    #     filename = file.filename.lower()

    #     try:
    #         if filename.endswith('.csv'):
    #             file_data = file.read().decode('utf-8')
    #             csv_reader = csv.DictReader(io.StringIO(file_data))
    #             for row in csv_reader:
    #                 leads.create(self._prepare_lead_vals(row))
    #                 # lead_vals = self._prepare_lead_vals(row)
    #                 # if lead_vals.get('lead_id'):
    #                 #     existing = leads.search([('lead_id', '=', lead_vals['lead_id'])], limit=1)
    #                 #     if existing:
    #                 #         return request.render('stag_elevators.import_error_template', {
    #                 #             'error': f"Duplicate Lead ID found: {lead_vals['lead_id']}"
    #                 #         })

    #                 # leads.create(lead_vals)

                    

    #         elif filename.endswith('.xlsx'):
    #             workbook = load_workbook(filename=io.BytesIO(file.read()))
    #             sheet = workbook.active
    #             headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    #             for row in sheet.iter_rows(min_row=2, values_only=True):
    #                 row_dict = dict(zip(headers, row))
    #                 leads.create(self._prepare_lead_vals(row_dict))
    #                 # lead_vals = self._prepare_lead_vals(row_dict)
    #                 # if lead_vals.get('lead_id'):
    #                 #     existing = leads.search([('lead_id', '=', lead_vals['lead_id'])], limit=1)
    #                 #     if existing:
    #                 #         return request.render('stag_elevators.import_error_template', {
    #                 #             'error': f"Duplicate Lead ID found: {lead_vals['lead_id']}"
    #                 #         })

    #                 # leads.create(lead_vals)
    #         else:
    #             # raise UserError(f"File not supported! {str(Exception)}")
    #             return request.render('stag_elevators.import_error_template', {
    #                 'error': f"File not supported: {str(Exception)}"
    #             })

    #     except Exception as e:
    #         # raise UserError(f"File not supported! {str(e)}")
    #         return request.render('stag_elevators.import_error_template', {
    #                 'error': f"File not supported!: {str(e)}"
    #             })

    #     return request.redirect('/opportunities')

    @http.route('/import/leads/submit', type='http', auth='user', website=True, csrf=False)
    def import_leads_submit(self, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'crm.lead'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })

        file = post.get('file')
        if not file:
            return request.render('stag_elevators.import_error_template', {
                'error': 'No file uploaded.'
            })

        leads = request.env['crm.lead'].sudo()
        filename = file.filename.lower()
        errors = []  # collect errors here
        imported_count = 0

        try:
            if filename.endswith('.csv'):
                file_data = file.read().decode('utf-8')
                csv_reader = csv.DictReader(io.StringIO(file_data))
                for idx, row in enumerate(csv_reader, start=2):  # row number for error reporting
                    try:
                        vals = self._prepare_lead_vals(row)
                        leads.create(vals)
                        imported_count += 1
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")

            elif filename.endswith('.xlsx'):
                workbook = load_workbook(filename=io.BytesIO(file.read()))
                sheet = workbook.active
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = dict(zip(headers, row))
                    try:
                        vals = self._prepare_lead_vals(row_dict)
                        leads.create(vals)
                        imported_count += 1
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
            else:
                return request.render('stag_elevators.import_error_template', {
                    'error': f"File not supported: {filename}"
                })

        except Exception as e:
            return request.render('stag_elevators.import_error_template', {
                'error': f"Unexpected error: {str(e)}"
            })

        # ✅ If there are errors, show error template with summary
        if errors:
            return request.render('stag_elevators.import_error_template', {
                'error': f"{imported_count} rows inserted below rows are skipped due to Errors found:",
                'errors': errors
            })

        # ✅ If no errors → success
        return request.redirect('/opportunities')

    def _prepare_lead_vals(self, row):
        name = (row.get('Name') or '').strip()
        phone = str(row.get('Number') or '').strip()
        email = (row.get('Email') or '').strip()

        if not name:
            raise Exception("Name is required")
        if not phone:
            raise Exception("Phone number is required")
        if phone and not (phone.isdigit() and 7 <= len(phone) <= 15):
            raise Exception(f"Invalid phone number: {phone}")
        if email:
            email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
            if not re.match(email_regex, email):
                raise Exception(f"Invalid email address: {email}")

        stage = request.env['crm.stage'].search([('name', '=', 'New')], limit=1)

        return {
            'lead_id': row.get('Lead ID'),
            'lead_platform': row.get('Lead Platform'),
            'name': name,
            'phone': phone,
            'email_from': email,
            'city': self._get_or_create_city(row.get('City/State')),
            'type_of_construction': row.get('Type of Construction'),
            'no_of_floor': row.get('No of Floors') or 0,
            'pre_comment': row.get('PreQualified Comment'),
            'user_id': self._get_user_id_by_name(row.get('Final ISM')),
            'stage_id': self._get_stage_id_by_name(row.get('Stages')) or stage.id,
            'prospect_hit': row.get('Prospect heat'),
            'lead_comment': row.get('Comments'),
        }

    # def _prepare_lead_vals(self, row):
    #     name = row.get('Name', '').strip()
    #     phone = str(row.get('Number', '') or '').strip()
    #     email = row.get('Email', '').strip()

    #     # Name validation
    #     if not name:
    #         raise UserError("Name is required for all leads.")

    #     # Phone validation (basic: digits and length between 7 to 15)
    #     if phone and not (phone.isdigit() and 7 <= len(phone) <= 15):
    #         raise UserError(f"Invalid phone number: {phone}")
        
    #     if not phone:
    #         raise UserError(f"Empty field Number")

    #     # Email validation using regex
    #     email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    #     if email and not re.match(email_regex, email):
    #         raise UserError(f"Invalid email address: {email}")
        
    #     stage = request.env['crm.stage'].search([('name', '=', 'New')], limit=1)
    #     return {
    #         'lead_id': row.get('Lead ID'),
    #         'lead_platform': row.get('Lead Platform'),
    #         'name': name,
    #         'phone': phone,
    #         'email_from': email,
    #         'city': self._get_or_create_city(row.get('City/State')),
    #         'type_of_construction': row.get('Type of Construction'),
    #         'no_of_floor': row.get('No of Floors') or 0,
    #         'pre_comment': row.get('PreQualified Comment'),
    #         'user_id': self._get_user_id_by_name(row.get('Final ISM')),
    #         'stage_id': self._get_stage_id_by_name(row.get('Stages')),
    #         'prospect_hit': row.get('Prospect heat'),
    #         'lead_comment': row.get('Comments'),
    #         'stage_id':stage.id,
    #     }
    
    def _get_or_create_city(self, city_name):
        if city_name:
            City = request.env['res.city'].sudo()
            city = City.search([('name', '=', city_name)], limit=1)
            if not city:
                city = City.create({'name': city_name})
            return city.name
        return False



    def _get_user_id_by_name(self, name):
        if name:
            user = request.env['res.users'].sudo().search([('login', '=', name)], limit=1)
            return user.name if user else False
        return False

    def _get_stage_id_by_name(self, stage_name):
        if stage_name:
            stage = request.env['crm.stage'].sudo().search([('name', '=', stage_name)], limit=1)
            return stage.name if stage else False
        return False


# def _prepare_lead_vals(row):
   
#     return {
#         'name': row.get('Name') or 'Unnamed Lead',
#         'email_from': row.get('Email'),
#         'phone': row.get('Number'),
#         'city': row.get('City/State'),
#         # 'description': row.get('Comments'),
#         'type': 'opportunity',
#         'lead_id': row.get('Lead ID'),
#         'lead_platform': row.get('Lead Platform'),
#         'type_of_construction': row.get('Type of Construction'),
#         'no_of_floor': row.get('No of Floors') or 0,
#         'pre_comment': row.get('PreQualified Comment'),
#         'prospect_hit': row.get('Prospect heat'),
#         'lead_comment': row.get('Comments'),
#         'user_id': row.get('ISM Name'),
        
#     }
    
class WebsiteQuotation(http.Controller):

    @http.route('/quotation/add', type='http', auth='user', website=True)
    def render_quotation_form(self, **kwargs):
        leads = request.env['crm.lead'].sudo().search([('email_from', '!=', False)])
        return request.render('stag_elevators.quotation_sale_form', {
            'leads': leads,
            'user': request.env.user.name,
            'today': fields.Date.today(),
        })

    @http.route('/quotation/submit', type='http', auth='user', methods=['POST'], website=True, csrf=False)
    def submit_quotation_form(self, **post):
        # Optional: create a contact first (if needed)
        partner = request.env['res.partner'].sudo().create({
            'name': post.get('name'),
            'phone': post.get('mobile'),
            'email': post.get('email'),
            'city': post.get('city'),
            


        })
        date_str = request.params.get('date_order')
        date_obj = datetime.strptime(date_str, '%Y-%m-%d') if date_str else fields.Datetime.now()
        # Create Sale Order
        request.env['sale.order'].sudo().create({
            'partner_id': partner.id,
            'date_order': date_obj,
            'note': f"Lead ID: {post.get('lead_id')}\nCreated By: {post.get('created_by')}",
            'installation':post.get('installation'),
            'travel_height':post.get('travel_height'),
            'pit_ramp':post.get('pit_ramp'),
            'minimum_headroom':post.get('minimum_headroom'),
            'capacity':post.get('capacity'),
            'motor':post.get('motor'),
            'speed':post.get('speed'),
            'power_supply':post.get('power_supply'),
            'power_absorption':post.get('power_absorption'),
            'cop':post.get('cop'),
            'cabin_size':post.get('cabin_size'),
            'flooring':post.get('flooring'),
            'cabin_walls':post.get('cabin_walls'),
            'cabin_panel':post.get('cabin_panel'),
            'required_shaft_space':post.get('required_shaft_space'),
            'no_of_floors':post.get('no_of_floors'),
            'amount_total':post.get('amount_total'),
            # 'commitment_date':post.get('commitment_date'),
            'warranty':post.get('warranty'),
        })

        return request.redirect('/quotation/list')
    
    
    
    @http.route('/quotation/list', type='http', auth='user', website=True)
    def quotation_list(self, **kwargs):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'ir.attachment'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        followups = request.env['lead.followup'].sudo().search([], order="next_followup_date asc")
        
        # Collect followup IDs
        followup_ids = followups.ids

        # Get all related attachments
        attachments = request.env['ir.attachment'].sudo().search([
            ('res_model', '=', 'lead.followup'),
            ('res_id', 'in', followup_ids),
        ])

        # Build attachment_map: {followup_id: {stage_label: attachment}}
        attachment_map = {}
        for att in attachments:
            fid = att.res_id
            # Default to 'Quotation' if no proper stage_label
            stage_label = 'Quotation'
            if att.description and 'Stage:' in att.description:
                stage_label = att.description.split('\n')[0].replace('Stage: ', '').strip()

            attachment_map.setdefault(fid, {})[stage_label] = att

        return request.render('stag_elevators.quotation_list_view', {
            'followups': followups,
            'attachment_map': attachment_map,  # 🔁 important line
        })
    
    @http.route('/quotation/view/<int:order_id>', type='http', auth='user', website=True)
    def quotation_view(self, order_id):
        order = request.env['sale.order'].sudo().browse(order_id)
        if not order:
            return request.not_found()
        return request.render('stag_elevators.quotation_view_template', {
            'order': order,
        })

    # Edit form
    @http.route('/quotation/edit/<int:order_id>', type='http', auth='user', website=True)
    def quotation_edit(self, order_id):
        order = request.env['sale.order'].sudo().browse(order_id)
        leads = request.env['crm.lead'].sudo().search([])
        if not order:
            return request.not_found()
        return request.render('stag_elevators.quotation_edit_form', {
            'order': order,
            'leads': leads,
            'user': request.env.user.name,
        })

    # Handle edit
    @http.route('/quotation/update/<int:order_id>', type='http', auth='user', methods=['POST'], website=True, csrf=True)
    def quotation_update(self, order_id, **post):
        order = request.env['sale.order'].sudo().browse(order_id)
        if order:
            order.write({
                'name': post.get('name'),
                'no_of_floors': post.get('no_of_floors'),
                'amount_total': post.get('amount_total'),
                'commitment_date': post.get('commitment_date'),
            })
        return request.redirect('/quotation/list')

    # Delete quotation
    @http.route('/quotation/delete/<int:order_id>', type='http', auth='user', website=True)
    def quotation_delete(self, order_id):
        order = request.env['sale.order'].sudo().browse(order_id)
        if order:
            order.unlink()
        return request.redirect('/quotation/list')

class LeadCalendarController(http.Controller):

    @http.route('/lead/calendar/events', type='http', auth='user', csrf=False)
    def lead_calendar_events(self):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        if my_groups.is_admin:
            leads = request.env['lead.followup'].sudo().search([('next_followup_date', '!=', False),('lead_id.stage_id.name', '!=', 'Won'),('lead_id.stage_id.name', '!=', 'Lost')])
            events = []

            for lead in leads:
                events.append({
                    'id': lead.id,
                    'title': lead.lead_id.name or 'No Title',
                    'start': lead.next_followup_date.strftime('%Y-%m-%d'),
                })

            return request.make_response(
                json.dumps(events),
                headers=[('Content-Type', 'application/json')])
        user = request.env.user
        leads = request.env['lead.followup'].sudo().search([('next_followup_date', '!=', False),('lead_id.user_id','=',user.id),('lead_id.stage_id.name', '!=', 'Won'),('lead_id.stage_id.name', '!=', 'Lost'),])
        events = []

        for lead in leads:
            events.append({
                'id': lead.id,
                'title': lead.lead_id.name or 'No Title',
                'start': lead.next_followup_date.strftime('%Y-%m-%d'),
            })

        return request.make_response(
            json.dumps(events),
            headers=[('Content-Type', 'application/json')]
    )
class AMCCalendarController(http.Controller):

    @http.route('/amc/calendar/events', type='http', auth='user', csrf=False)
    def amc_calendar_events(self):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        if my_groups.is_admin:
            leads = request.env['crm.client'].sudo().search([('renew_date', '!=', False) and ('renew','=',True) ])
            crm_followups = request.env['crm.client'].sudo().search([('date','!=',False)])
            events = []

            for lead in leads:
                events.append({
                    'id': lead.id,
                    'title': lead.lead_id.name or 'No Title',
                    'start': lead.renew_date.strftime('%Y-%m-%d'),
                    'color': 'red',
                })
            
            for followup in crm_followups:
                events.append({
                    'id': followup.id,
                    'title': followup.lead_id.name or 'No Title',
                    'start': followup.date.strftime('%Y-%m-%d'),
                    
                })

            return request.make_response(
                json.dumps(events),
                headers=[('Content-Type', 'application/json')]
            )
        else:
            # leads = request.env['crm.client'].sudo().search([('renew_date', '!=', False) and ('renew','=',True) ])
            leads = request.env['crm.client'].sudo().search([
                ('renew_date', '!=', False),
                ('renew', '=', True),
                ('assign_to', '=', request.env.user.id)
            ])
            prod_lead = request.env['production.client'].sudo().search([('crm_client_id.renew_date','!=',False)])
            crm_followups = request.env['crm.client'].sudo().search([('date','!=',False),('assign_to', '=', request.env.user.id)])
            events = []

            for lead in leads:
                events.append({
                    'id': lead.id,
                    'title': lead.lead_id.name or 'No Title',
                    'start': lead.renew_date.strftime('%Y-%m-%d'),
                    'color': 'red',
                })
            
            for followup in crm_followups:
                events.append({
                    'id': followup.id,
                    'title': followup.lead_id.name or 'No Title',
                    'start': followup.date.strftime('%Y-%m-%d'),
                    
                })
            
            for followup in prod_lead:
                events.append({
                    'id': followup.id,
                    'title': followup.crm_client_id.lead_id.name or 'No Title',
                    'start': followup.crm_client_id.renew_date.strftime('%Y-%m-%d'),
                    'url': '/production/client/edit/%s' % followup.id,
                    'color':'red'
                })

            return request.make_response(
                json.dumps(events),
                headers=[('Content-Type', 'application/json')]
            )
        
    
    @http.route('/production/calendar/events', type='http', auth='user', csrf=False)
    def production_calendar_events(self):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        if my_groups.is_admin:
            leads = request.env['production.client'].sudo().search([('material_dispatch_date','!=',False)])
            production_followups = request.env['production.client'].sudo().search([('followup_date','!=',False)])
            events = []

            for lead in leads:
                events.append({
                    'id': lead.id,
                    'title': lead.lead_id.name or 'No Title',
                    'start': lead.material_dispatch_date.strftime('%Y-%m-%d'),
                    'color':'red',
                })
            if production_followups:    
                for lead in production_followups:
                    events.append({
                        'id': lead.id,
                        'title': lead.lead_id.name or 'No Title',
                        'start': lead.followup_date.strftime('%Y-%m-%d'),
                    })
            
            

            return request.make_response(
                json.dumps(events),
                headers=[('Content-Type', 'application/json')]
            )
        else:
            # leads = request.env['crm.client'].sudo().search([('renew_date', '!=', False) and ('renew','=',True) ])
            leads = request.env['production.client'].sudo().search([
                ('assign_to','=',request.env.user.id)
            ])
            production_followups = request.env['production.client'].sudo().search([('followup_date','!=',False),('assign_to','=',request.env.user.id)])
            events = []

            for lead in leads:
                events.append({
                    'id': lead.id,
                    'title': lead.lead_id.name or 'No Title',
                    'start': lead.material_dispatch_date.strftime('%Y-%m-%d'),
                    'color':'red'
                })
            if production_followups:
                for lead in production_followups:
                    events.append({
                        'id': lead.id,
                        'title': lead.lead_id.name or 'No Title',
                        'start': lead.followup_date.strftime('%Y-%m-%d'),
                    })
            
            

            return request.make_response(
                json.dumps(events),
                headers=[('Content-Type', 'application/json')]
            )



class UserPermissionController(http.Controller):

    @http.route('/user_permissions/save', type='http', auth='user', methods=['POST'], csrf=False)
    def save_user_permissions(self, **post):
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'ir.model.access'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.render('stag_elevators.access_denied_template', {
                'message': 'You do not have permission to view.'
            })
        role_id = int(post.get('role_id'))
        group = request.env['res.groups'].sudo().browse(role_id)

        if not group:
            return request.redirect('/user_permissions')  # fallback if group not found

        # Example: update access based on form input
        permissions = {
            'user_role': 'res.groups',        # model_name
            'city_creation': 'res.city', # model_name
            'user_permissions':'ir.model.access',
            'user_creation':'res.users',
            'lead_stage':'crm.stage',
            'quotation':'ir.attachment',
            'opportunity_creation':'crm.lead',
            'lead_followup':'lead.followup',
            'crm':'crm.client',
            'production':'production.client',
            
        }
        if group.exists():
            for field, model_xml in permissions.items():
                value = post.get(field)
                model = request.env['ir.model'].sudo().search([('model', '=', model_xml)], limit=1)

                if model:
                    # Get or create the record in ir.model.access
                    access = request.env['ir.model.access'].sudo().search([
                        ('group_id', '=', group.id),
                        ('model_id', '=', model.id)
                    ], limit=1)
                    if not access:
                        access = request.env['ir.model.access'].sudo().create({
                            'name': f'{group.name} access for {model.model}',
                            'model_id': model.id,
                            'group_id': group.id,
                            'perm_read': value == 'Yes',
                            'perm_write': value == 'Yes',
                            'perm_create': value == 'Yes',
                            'perm_unlink': value == 'Yes',
                        })
                    else:
                        access.write({
                            'perm_read': value == 'Yes',
                            'perm_write': value == 'Yes',
                            'perm_create': value == 'Yes',
                            'perm_unlink': value == 'Yes',
                        })

        return request.redirect('/user_permissions')
    
class UserController(http.Controller):

    @http.route('/users/impersonate/<int:user_id>', type='http', auth='user', website=True)
    def impersonate_user(self, user_id, **kwargs):
        # Only allow admin
        my_groups = request.env.user.groups_id.filtered(lambda g: g.ismy)
        access_rule = request.env['ir.model.access'].sudo().search([
                ('model_id.model', '=', 'ir.model.access'),
                ('group_id', '=', my_groups.id),
                ('perm_read', '=', True)
            ], limit=1)
        if not access_rule:
            return request.redirect('/Dashboard')

        # Save current admin UID if not already saved
        if 'original_uid' not in request.session:
            request.session['original_uid'] = request.session.uid

        # Impersonate target user
        user = request.env['res.users'].sudo().browse(user_id)
        if not user.exists():
            return request.redirect('/Dashboard')

        request.session.uid = user.id
        return request.redirect('/Dashboard')

    @http.route('/users/stop_impersonate', type='http', auth='user', website=True)
    def stop_impersonate(self, **kwargs):
        # Restore admin session if impersonating
        if 'original_uid' in request.session:
            request.session.uid = request.session.pop('original_uid')
        return request.redirect('/Dashboard')
    

    @http.route('/my/change_password', type='http', auth='user', methods=['POST'], csrf=False)
    def change_password(self, **kwargs):
        data = json.loads(request.httprequest.data.decode())


        new_password = data.get("new_password")
        confirm_password = data.get("confirm_password")

        user = request.env.user

        if new_password != confirm_password:
            return request.make_response(json.dumps({"success": False, "message": "Passwords do not match."}),
                                         headers=[("Content-Type", "application/json")])

        user.sudo().write({'password': new_password,'signature':new_password})

        return request.make_response(json.dumps({"success": True, "message": "Password changed successfully!"}),
                                     headers=[("Content-Type", "application/json")])

